import time
import smtplib
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

def sendmail(bcc):
    gmail_user = "itsybitsyfinance@gmail.com"
    gmail_password = "mohit69tushar"

    sent_from = gmail_user
    to = ["das.5@iitj.ac.in"]
    cc = ["babra.1@iitj.ac.in","suniketdas07011999@gmail.com","agarwal.6@iitj.ac.in"]
    bcc = ["mukulkumar258@gmail.com"]
    subject = "Lorem ipsum dolor sit amet"
    body = "consectetur adipiscing elit"


    email_text = """\
    From: %s
    To: %s
    CC: %s
    Subject: %s

    %s
    """ % (sent_from, ", ".join(to), ", ".join(cc), subject, body)

    toaddrs = [to] + cc + bcc

    try:
        smtp_server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        smtp_server.ehlo()
        smtp_server.login(gmail_user, gmail_password)
        smtp_server.sendmail(sent_from, toaddrs, email_text)
        smtp_server.close()
        print ("Email sent successfully!")
    except Exception as ex:
        print ("Something went wrong….",ex)

def addDirect(arr,k):
    wb = load_workbook(r"C:\Users\Mukul\Desktop\Github\Itsy Bitsy Finance\Stocks.xlsx")
    sheet = wb.active
    j=2
    for i in range (0,940):
        sheet.cell(row=j, column=k).value=arr[i]
        j+=1
    wb.save(r"C:\Users\Mukul\Desktop\Github\Itsy Bitsy Finance\Stocks.xlsx")   
    wb.close() 

def addYesNo(dma,k):
    wb = load_workbook(r"C:\Users\Mukul\Desktop\Github\Itsy Bitsy Finance\Stocks.xlsx")
    sheet = wb.active
    i=0
    j=2
    while(j<=941):
        sheet.cell(row=j, column=k).value="No"
        j=j+1 
     
    while(i<len(dma)):
        j=2   
        while(j<941):
            if (dma[i]==sheet.cell(row=j, column=1).value):
                sheet.cell(row=j, column=k).value="Yes"
            j=j+1 
        i=i+1
    wb.save(r"C:\Users\Mukul\Desktop\Github\Itsy Bitsy Finance\Stocks.xlsx")   
    wb.close()  

def open_site(driver):
    driver.get('https://in.tradingview.com/screener/') 

def maximize(driver):
    driver.maximize_window()

def getstocks(driver):
    i=1
    allstocks=[]
    time.sleep(5)
    while(i<=gettickernum(driver)):
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[1]/div/div[2]/a")
        stockname=element.text
        allstocks.append(stockname)
        i=i+1
        if(i%150==0):
            scrollbottom(driver)
    return allstocks      

def getfunda(driver):
    i=1
    allstocks=[]
    pe=[]
    mktcap=[]
    vol=[]
    curprice=[]
    time.sleep(5)
    while(i<=941):
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[1]/div/div[2]/a")
        temp=element.text
        allstocks.append(temp)
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[8]")
        temp=element.text
        pe.append(temp)
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[7]")
        temp=element.text
        mktcap.append(temp)
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[6]")
        temp=element.text
        vol.append(temp)
        element =driver.find_element(By.XPATH,"/html/body/div[3]/div/div[4]/table/tbody/tr["+str(i)+"]/td[2]/span[1]")
        temp=element.text
        curprice.append(temp)
        i=i+1
        if(i%150==0):
            scrollbottom(driver)
    return allstocks,mktcap,vol,pe,curprice

def containsNumber(value):
    str=""
    for character in value:
        if character.isdigit():
            str=str+character
    str=int(str)
    return str

def gettickernum(driver):
    element= driver.find_element(By.XPATH,"/html/body/div[3]/div/div[3]/table/thead/tr/th[1]/div/div/div[2]").text
    tickernum=containsNumber(element)
    return tickernum


def scrollbottom(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")  
    time.sleep(5)
    

def clickfilter(driver):
    time.sleep(1)
    element=driver.find_element(By.XPATH,"/html/body/div[3]/div/div[2]/div[14]") 
    element.click()

def searchfilter(driver,filtername):
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[2]/input")
    element.send_keys(filtername)

def closefilter(driver):
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[1]/div[3]") 
    element.click()

def fromMktcap(driver,percentincrease):
    clickfilter(driver)
    searchfilter(driver,"Market Capitalization")
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[3]/div[1]/div/div/div[8]/div[2]/div/span[1]")
    element.click()
    i=0
    while(i<(percentincrease/10)):
        element.send_keys(Keys.RIGHT, Keys.RETURN)
        i=i+1
    closefilter(driver)


def DMAabove(driver,filtername,filternum):
    clickfilter(driver)
    searchfilter(driver,filtername)
    time.sleep(2)
    #dropdown menu
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[3]/div[1]/div/div/div["+str(filternum)+"]/div[2]/label[1]/span/i")
    element.click()
    #select above
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[4]/div/div[1]/span[3]")
    element.click()
    #dropdown menu
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[3]/div[1]/div/div/div["+str(filternum)+"]/div[2]/label[2]/span/i")
    element.click()
    #select last
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[4]/div/div[1]/span[5]")
    element.click()
    closefilter(driver)
    stonks=getstocks(driver)
    clickfilter(driver)
    element=driver.find_element(By.XPATH,"/html/body/div[8]/div/div/div/div[3]/div[1]/div/div/div[4]/div[3]")
    element.click()
    closefilter(driver) 
    return stonks


if __name__=="__main__":
    webdriverpath="D:\software\chrome webdriver\chromedriver.exe"
    driver = webdriver.Chrome(webdriverpath)
    open_site(driver)
    fromMktcap(driver,50)
    allstocks,mktcap,vol,pe,curprice =getfunda(driver) 
    addDirect(allstocks,1)
    addDirect(curprice,2)
    addDirect(mktcap,3)
    addDirect(vol,4)
    addDirect(pe,5)
    stocks=DMAabove(driver,"Simple Moving Average (200)",91)
    addYesNo(stocks,6)
    stocks=DMAabove(driver,"Simple Moving Average (100)",90)
    addYesNo(stocks,7)
    stocks=DMAabove(driver,"Simple Moving Average (50)",89)
    addYesNo(stocks,8)
    